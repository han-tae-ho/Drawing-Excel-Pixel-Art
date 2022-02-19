from PIL import Image
import numpy as np
import matplotlib.pyplot as plt
import cv2
import openpyxl as xl
import timm
from openpyxl.styles import Color, PatternFill
from tqdm import tqdm
import argparse 

parser = argparse.ArgumentParser(description='Drawing Excel Pixel Art using Python !!') 
parser.add_argument('-W', '--width', default=50, type=int, help='set width') 
parser.add_argument('-H', '--height', default=50, type=int, help='set height') 
parser.add_argument('--path', default='./sample.jpeg', help= 'input image path') 
args = parser.parse_args()

# set width, height, image_path
W = args.width
H = args.height
img_path = args.path
def rgb_to_hex(arr):
    r, g, b, a = arr
    return  hex(a)[2:].zfill(2) + hex(r)[2:].zfill(2) + hex(g)[2:].zfill(2) + hex(b)[2:].zfill(2)

def img2exl():
    img = Image.open(img_path) # put image path. (file - jpg, png, jpeg ...)
    img = img.resize((W, H))
    img = np.array(img)

    # PatternFill fuction needs aRGB hex prameter.
    # so, if input image has not 4 channel, covert it into 4 channel.
    if len(img.shape) == 3:
        #convert the image from RGBA2RGB
        img = cv2.cvtColor(img, cv2.COLOR_RGB2RGBA)
    elif len(img.shage) == 2:
        img = cv2.cvtColor(img, cv2.COLOR_GRAY2RGBA)



    pixels = img.flatten().reshape(-1, 4)

    wb = xl.load_workbook('./blueprint.xlsx')
    ws = wb.active

    i = 0
    for row_num in tqdm(range(1,H + 1)): 
        for col_num in range(1,W + 1):
            hex_code = rgb_to_hex(pixels[i])
            Fill = PatternFill(start_color=hex_code,
                       end_color=hex_code,
                       fill_type='solid')
            select_cell = ws.cell(row=row_num, column=col_num)
            select_cell.fill = Fill
            i += 1

    wb.save('output.xlsx')

if __name__ == "__main__":
    img2exl()